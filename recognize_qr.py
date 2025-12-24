import os
import time
from openpyxl import Workbook, load_workbook
import cv2

# ========== Excel setup ==========
base_path = os.path.dirname(os.path.abspath(__file__))
excel_path = os.path.join(base_path, "students_db.xlsx")

students = [
    {"ID": 202501, "Name": "Aditya"},
    {"ID": 202502, "Name": "Rahul"},
    {"ID": 202503, "Name": "Sneha"},
]

# Create Excel if it doesn't exist
if not os.path.exists(excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["ID", "Name", "Status"])

    for i, student in enumerate(students, start=2):
        ws[f"A{i}"] = student["ID"]
        ws[f"B{i}"] = student["Name"]
        ws[f"C{i}"] = "Absent"

    wb.save(excel_path)
    print("✅ Excel created with student list.")
else:
    print("📘 Using existing Excel file.")

# ========== Attendance marking ==========
def mark_present(student_id):
    wb2 = load_workbook(excel_path)
    ws2 = wb2.active
    found = False
    message = ""

    for row in ws2.iter_rows(min_row=2, values_only=False):
        cell_id = row[0].value
        cell_status = row[2]
        if cell_id == int(student_id):
            if cell_status.value != "Present":
                cell_status.value = "Present"
                message = f"✅ Present marked for ID: {student_id}"
            else:
                message = f"⚠️ ID {student_id} already marked Present."
            found = True
            break

    if not found:
        message = f"❌ ID {student_id} not found in Excel!"

    wb2.save(excel_path)
    wb2.close()
    print(message)
    return message

# ========== QR detection ==========
cap = cv2.VideoCapture(0)
detector = cv2.QRCodeDetector()
scanned_ids = set()
last_message = ""
last_time = 0

print("🎥 QR Attendance System Started. Press 'q' to quit.\n")

while True:
    ret, frame = cap.read()
    if not ret:
        print("⚠️ Cannot read from webcam.")
        break

    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    data, bbox, _ = detector.detectAndDecode(gray)

    if bbox is not None:
        pts = bbox.astype(int).squeeze()
        if pts.ndim == 2:
            for i in range(len(pts)):
                pt1 = tuple(pts[i])
                pt2 = tuple(pts[(i + 1) % len(pts)])
                cv2.line(frame, pt1, pt2, (0, 255, 0), 2)

    if data:
        student_id = data.strip()
        if student_id not in scanned_ids:
            last_message = mark_present(student_id)
            last_time = time.time()
            scanned_ids.add(student_id)

        cv2.putText(frame, f"ID: {student_id}", (30, 50),
                    cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)

    # Show “Present marked” message for 3 seconds
    if last_message and (time.time() - last_time) < 3:
        cv2.putText(frame, last_message, (30, 100),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)

    cv2.imshow("📷 QR Attendance System", frame)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()
print("\n📁 Attendance saved in 'students_db.xlsx'")
